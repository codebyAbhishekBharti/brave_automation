"""
This is a module for calculating the bat earned for every profile and store it into excel file.
1. open the brave browser
2. Open new tab
3. Select the data area
4. Copy the data
5. Paste the data into excel file and also make min and max to just directly show the data
6. Close the tab
7. repeat the process for all the profiles
Improve the way of data storage as we have to transpose the data to store it into excel file
"""
import os
import time
import pyautogui
import threading
import keyboard
import pyperclip
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook


class Automate_brave_bat_calc:
    def __init__(self,total_profiles,total_adds=6,starting_profile=1):
        global thread_work_completed  # Use the global keyword to modify the global variable
        self.m=0
        self.M=0
        if(total_profiles<1):
            print("Invalid total profiles, total profiles must be greater than 0")
            thread_work_completed=True
            return None
        else:
            self.total_profiles=total_profiles                        #Sets total profiles for automation
        self.screen_width, self.screen_height = pyautogui.size()  #gets the screen size
        self.total_adds=total_adds                                #set to 7, 6 is exact
        if(starting_profile<1 or starting_profile>total_profiles):#checks if starting profile is valid or not
            print("Invalid starting profile, starting profile must be 1<=starting_profile<=total_profiles")
            thread_work_completed=True
            return None
        else :
            self.starting_profile=starting_profile                  #sets the starting profile
        self.excel_data=[] #Date Time Profile Min Max
        self.browser_opener()
        self.sotre_excel_data()
        print("Minimum",round(self.m,3))
        print("Maximum",round(self.M,3))
        thread_work_completed=True
    
    def sotre_excel_data(self):
        """This func will store the excel data into excel file"""
        excel_file_path = "data.xlsx"                # Specify the Excel file path
        try:
            # Try to load the existing workbook
            wb = load_workbook(excel_file_path)
            ws = wb.active
            for row in self.excel_data:                   # Insert data
                ws.append(row)
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook, add header, and insert data
            wb = Workbook()
            ws = wb.active
            header = ["Date", "Time", "Profile", "Min", "Max"]  # Add header
            ws.append(header)
            for row in self.excel_data:                    # Insert data
                ws.append(row)
        finally:
            wb.save(excel_file_path)                       #save the workbook
            print(f"Data has been successfully written to {excel_file_path}")
    
    def browser_opener(self):
        """this func will handle the browser opener"""
        os.system("start brave")              #starts edge broswer
        time.sleep(2)                         #wats for 2 seconds so that browser can open successfully
        pyautogui.press('tab')                #presses tab to select profile
        pyautogui.press('enter')              #presses enter to open profile
        if(self.starting_profile==1):
            print(f"Profile 1")              
            self.excel_data.append([f"Profile 1"])
            self.automator()                      #runs automator method to start automation works
            time.sleep(0.2)                       #wait a bit so that browser is ready to close
            self.starting_profile+=1

        for i in range(self.starting_profile,self.total_profiles+1):      #loops through the total profiles
            if(i in [12,13,14,15,17,19]):           #skips the profiles which are not working
                continue
            print(f"Profile {i}")
            time.sleep(1)                         #wait a bit so that browser is ready to close 
            pyautogui.hotkey('ctrl','shift','m')  #opens profile view
            time.sleep(0.2)                       #wait a bit so that browser is ready to close
            for j in range(i-2):
                pyautogui.press('tab')
            pyautogui.press('enter')
            time.sleep(0.2)                       #wait a bit so that browser is ready to close
            self.excel_data.append([f"Profile {i}"])
            self.automator()                      #runs automator method to start automation works
            pyautogui.hotkey('alt', 'f4')         #closing profiles to save RAM from getting full
        pyautogui.hotkey('alt', 'f4')             #closing main profile

    def window_centre_click(self):
        """this func will find the centre of the page and click at the centre"""
        x = self.screen_width // 2       #finds the centre of the page
        y = self.screen_height // 2      #finds the centre of the page
        pyautogui.click(x, y)            # Click on the center

    def new_tab(self):
        """This func will open new tab properly """
        time.sleep(0.5)                 #sets delay in pressing tabs so that browser can load properly
        pyautogui.hotkey('ctrl','t')    #opens new tab
        time.sleep(1)                   #sets delay so that new tab can open properly
        self.window_centre_click()      #clicks on the centre of the page so that page view can be worked on not the address bar

    def data_area_selecter(self):
        """this func will select the data area and return the selected data"""
        start_x, start_y , end_x, end_y = (1590, 640, 1845, 640) #sets the coordinates of the data area
        pyautogui.mouseDown(start_x, start_y)        # Simulate a mouse click to start the selection
        pyautogui.moveTo(end_x, end_y, duration=0)   # Move the mouse cursor to selection area and adjust duration as needed
        pyautogui.mouseUp()                          # Release the mouse
        pyautogui.hotkey('ctrl', 'c')                #copy the selected text
        copied_data = pyperclip.paste()              #paste the copied data into a variable
        return copied_data

    def data_storer(self,copied_data):
        """This func is will show total profit of the day and store data in excel file"""
        try:
            numeric_values = [float(match) for match in re.findall(r'[-+]?\d*\.\d+|\d+', copied_data)]
        except:
            print("No data found")
            numeric_values = [0,0]
        min_value = numeric_values[0]
        max_value = numeric_values[1]
        self.m+=min_value
        self.M+=max_value
        current_datetime = datetime.now()
        current_date = current_datetime.strftime("%d/%m/%Y")
        current_time = current_datetime.strftime("%H:%M:%S")
        self.excel_data[len(self.excel_data)-1].insert(0,current_time)
        self.excel_data[len(self.excel_data)-1].insert(0,current_date)
        self.excel_data[len(self.excel_data)-1].append(min_value)
        self.excel_data[len(self.excel_data)-1].append(max_value)

    def automator(self):
        """this function will handle all the process to automate particular page"""
        time.sleep(0.7)                             #sets delay in pressing tabs so that browser can load properly
        self.new_tab()                              #opens new tab
        copied_data = self.data_area_selecter()     #selects the data area
        self.data_storer(copied_data)               #copies and paste the data into excel file
        pyautogui.hotkey('ctrl','w')                #closes the add view tab

def program_terimator():
    """Terminates the whole program if 'q' is pressed or Ctrl+C is used."""
    try:
        while True:
            if keyboard.is_pressed('q') or thread_work_completed:            #checks it 'q' is pressed
                print("Jai shree ram")
                exit()                              #terminates the whole program
            time.sleep(0.1)                         #sets delay so that cpu can rest a bit
    except KeyboardInterrupt:
        print("\nCtrl+C detected. Terminating the program.")
        exit()                                      #terminates the whole program

if __name__ == '__main__':
    total_profiles=24
    total_adds=7
    starting_profile=1                #it must be 1<=starting_profile<=total_profiles
    thread_work_completed = False     # Initialize the global variable
    automation_thread = threading.Thread(target=Automate_brave_bat_calc, args=(total_profiles,total_adds,starting_profile,),daemon=True) #daemon is set true so that program can be terminated by pressing 'q'
    automation_thread.start()         #starts the thread
    program_terimator()               #starts the program terminator