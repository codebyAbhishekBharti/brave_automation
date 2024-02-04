"""
The automation program will help me to automate brave to earn bat
1. Open brave
2. Open first profile
3. click new tab
3. scroll down
4. click on the first news
5. Go to new tab
6. scroll down till add
7. refresh page
8. do refresh and scroll 6 times
9. switch to new account and do all again
"""

import pyautogui
import time
import os
import cv2
import threading
import keyboard
import numpy as np
from datetime import datetime
import pyperclip
import re
from openpyxl import Workbook, load_workbook

class Automate_brave:
    def __init__(self,total_profiles,total_adds=6,starting_profile=1,software_mode=1):
        #remove failsafe of pyautogui
        pyautogui.FAILSAFE = False
        if(total_profiles<1):
            print("Invalid total profiles, total profiles must be greater than 0")
            exit()
        else:
            self.total_profiles=total_profiles                        #Sets total profiles for automation
        self.screen_width, self.screen_height = pyautogui.size()  #gets the screen size
        self.total_adds=total_adds                                #set to 7, 6 is exact

        if(starting_profile<1 or starting_profile>total_profiles):#checks if starting profile is valid or not
            print("Invalid starting profile, starting profile must be 1<=starting_profile<=total_profiles")
            exit()
        else :
            self.starting_profile=starting_profile                  #sets the starting profile
        self.excel_data=[]                                    #Date Time Profile Min_start Max_satrt Min_end Max_end Total_Adds
        self.software_mode=software_mode
        self.directory_path = "D:\\mycodes\\python codes\\brave auto\\"
        if(self.software_mode!=1):
            self.excel_file_path = f"{self.directory_path}new_data.xlsx"                # Specify the Excel file path
            self.excel_file_opener()                            #opens the excel file
        self.browser_opener()
        if(self.software_mode!=1):
            print(f"Data has been successfully written to {self.excel_file_path}")

    def excel_file_opener(self):
        """This func will open the excel file"""
        try:
            # Try to load the existing workbook
            self.wb = load_workbook(self.excel_file_path)
            self.ws = self.wb.active
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook, add header, and insert data
            self.wb = Workbook()
            self.ws = self.wb.active
            header = ["Date", "Time", "Profile", "Min_start", "Max_start", "Min_end", "Max_end", "Total_Adds","Min_earn","Max_earn"]  # Add header
            self.ws.append(header)
            try:
                self.wb.save(self.excel_file_path)                                         #saves the data to the excel file
            except Exception as e:
                print("Error: ",e,"\nYou might have currently opened excel file")
                exit()

    def browser_opener(self):
        """this func will handle the browser opener"""
        os.system("start brave")              #starts edge broswer
        time.sleep(2)                         #wats for 2 seconds so that browser can open successfully
        pyautogui.press('tab')                #presses tab to select profile
        pyautogui.press('enter')              #presses enter to open profile
        if(self.starting_profile==1):
            print(f"Profile 1")              
            self.profile_no=1
            self.automator()                      #runs automator method to start automation works
            time.sleep(0.2)                       #wait a bit so that browser is ready to close
            self.starting_profile+=1

        for i in range(self.starting_profile,self.total_profiles+1):      #loops through the total profiles
            if(i in [12,13,14,15,17,19]):           #skips the profiles which are not working
                continue
            print(f"Profile {i}")
            self.profile_no=i
            time.sleep(1)                         #wait a bit so that browser is ready to close 
            pyautogui.hotkey('ctrl','shift','m')  #opens profile view
            time.sleep(0.2)                       #wait a bit so that browser is ready to close
            for j in range(i-2):
                pyautogui.press('tab')
            pyautogui.press('enter')
            time.sleep(0.2)                       #wait a bit so that browser is ready to close
            self.automator()                      #runs automator method to start automation works
            pyautogui.hotkey('alt', 'f4')         #closing profiles to save RAM from getting full
        # pyautogui.hotkey('alt', 'f4')             #closing main profile

    def new_tab(self):
        """This func will open new tab properly """
        time.sleep(0.5)                 #sets delay in pressing tabs so that browser can load properly
        pyautogui.hotkey('ctrl','t')    #opens new tab
        time.sleep(1)                   #sets delay so that new tab can open properly
        pyautogui.click(128,205)        #Random click to get out of address bar
    
    def take_screenshot(self):
        """This func will take screenshot and return it"""
        screenshot = pyautogui.screenshot()
        screenshot = np.array(screenshot)
        screenshot = cv2.cvtColor(screenshot, cv2.COLOR_RGB2BGR)
        return screenshot

    def check_image_presence(self,template_path):
        """This func will check if the template image is present in the screenshot"""
        template = cv2.imread(template_path)
        screenshot = self.take_screenshot()
        result = cv2.matchTemplate(screenshot, template, cv2.TM_CCOEFF_NORMED)
        threshold = 0.8  # Adjust this threshold based on your needs
        locations = np.where(result >= threshold)
        return len(locations[0]) > 0

    def current_bat_data(self):
        """this func will select the data area and return  min and max bat data"""
        start_x, start_y , end_x, end_y = (1590, 640, 1845, 640) #sets the coordinates of the data area
        time.sleep(1)                              #sets delay so that page can load properly
        pyautogui.mouseDown(start_x, start_y)        # Simulate a mouse click to start the selection
        pyautogui.moveTo(end_x, end_y, duration=0)   # Move the mouse cursor to selection area and adjust duration as needed
        pyautogui.mouseUp()                          # Release the mouse
        pyautogui.hotkey('ctrl', 'c')                #copy the selected text
        copied_data = pyperclip.paste()              #paste the copied data into a variable
        try:
            numeric_values = [float(match) for match in re.findall(r'[-+]?\d*\.\d+|\d+', copied_data)]
            if(len(numeric_values)!=2):
                print("Data is not found")
                numeric_values = [0,0]
        except:
            print("No data found")
            numeric_values = [0,0]
        return numeric_values
    
    def excel_data_storer(self,data_type=1,total_adds_viwed=0):
        """This func will store the data in excel
            data_type=1 for storing starting bat data
            data_type=2 for storing ending bat data
        """
        if(data_type<1 or data_type>2):
            raise ValueError("Invalid data type, data type must be \n1. Starting data storage \n 2. Ending data storage")
        current_data = self.current_bat_data()           #gets the current bat data
        if(data_type==1):                                  #if data type is 1 then it will store the starting bat data
            current_datetime = datetime.now()               #gets the current date and time
            current_date = current_datetime.strftime("%d/%m/%Y")  #gets the current date
            current_time = current_datetime.strftime("%H:%M:%S")  #gets the current time
            self.excel_data.extend([current_date,current_time,f"profile {self.profile_no}",current_data[0],current_data[1]])  #appends the date time and profile number
        else:                                                #if data type is 2 then it will store the ending bat data
            self.excel_data.append(current_data[0])                                #appends the ending bat data
            self.excel_data.append(current_data[1])                                #appends the ending bat data
            self.excel_data.append(total_adds_viwed)                               #appends the total adds viewed
            self.excel_data.append(current_data[0]-self.excel_data[3])      #appends the total profit
            self.excel_data.append(current_data[1]-self.excel_data[4])      #appends the total profit
            self.ws.append(self.excel_data)                                       #appends the data to the excel file
            self.excel_data=[]                                                     #clears the data for next profile
            try:
                self.wb.save(self.excel_file_path)                                         #saves the data to the excel file
            except Exception as e:
                print("Error: ",e,"\nYou might have currently opened excel file")
                exit()

    def wait_for_news_load(self):
        """This func will wait for the news to load properly"""
        # self.news_activator()
        for i in range(3):                                   #loops 3 chances to make sure news is loaded properly
            self.new_tab()                                   #opens new tab
            if(self.software_mode!=1 and i==0):              #if software mode is not 1 then only it will store the data
                self.excel_data_storer(1)                    #stores the starting bat data
                if(self.software_mode==2): return            #if software mode is 2 then it will not open the news
            for _ in range(7):                               #for every loop wait for 7 seconds to show the news
                if(self.check_image_presence(f"{self.directory_path}news.png")):   #checks if news is present or not
                    time.sleep(1)
                    pyautogui.click(945,840)                 #clicks on the news icon
                    time.sleep(1.2)
                    return 1
                time.sleep(0.6)
            pyautogui.hotkey('ctrl','w')                     #closes the news view tab
        return -1
    
    def add_view_handler(self):
        """this func will make sure to view all adds """
        x = self.screen_width // 2              #finds the centre of the page on x axis
        viwed_adds=0                            #sets the counter for adds viewed
        if(self.software_mode!=2):             #if software mode is not 2 then only it will store the data
            for i in range(self.total_adds):        #loops through the total adds
                time.sleep(1)
                pyautogui.scroll(-1000)         #scrolls down to add
                time.sleep(1)                   #sets delay so that page can load properly
                if(not self.check_image_presence(f"{self.directory_path}Ad_logo_new.png")):       #checks if add is present or not\
                    break
                viwed_adds+=1
                pyautogui.press('f5')               #refreshes the page for new add
        if(self.software_mode!=1):                  #if software mode is not 1 then only it will store the data
            for _ in range(4):
                pyautogui.press('pageup')         #scrolls up to the top of the page
            pyautogui.press('f5')               #refreshes the page for new add
            self.excel_data_storer(2,viwed_adds)  #stores the ending bat data
        pyautogui.hotkey('ctrl','w')            #closes the add view tab
        time.sleep(0.1)                         #waits till current tab closes

    def automator(self):
        """this function will handle all the process to automate particular page"""
        time.sleep(0.7)             #sets delay in pressing tabs so that browser can load properly
        self.wait_for_news_load()      #opens first news
        self.add_view_handler()     #handles all the adds to view properly

def program_terminator():
    """Terminates the whole program if 'q' is pressed or Ctrl+C is used."""
    try:
        while True:
            if keyboard.is_pressed('q'):
                print("Jai shree ram")
                os._exit(0)  # Provide an exit code
            time.sleep(0.1)
    except KeyboardInterrupt:
        print("\nCtrl+C detected. Terminating the program.")
        os._exit(0)

if __name__ == '__main__':
    total_profiles=24
    total_adds=7
    starting_profile=1                #it must be 1<=starting_profile<=total_profiles

    software_mode = 3                 # 1 for Adds only, 2 for calculuator only, 3 for both

    terinator_thread = threading.Thread(target=program_terminator,args=(),daemon=True)
    terinator_thread.start()
    automation_thread = Automate_brave(total_profiles,total_adds,starting_profile,software_mode)